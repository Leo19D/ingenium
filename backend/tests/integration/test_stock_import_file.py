"""Integration: uvoz skladišta iz fajla (Excel + PDF) — /stock-items/import-file."""

from __future__ import annotations

import io
import uuid

import pytest
import pytest_asyncio
from fpdf import FPDF
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.api.deps import get_current_org_id
from app.db.models.organization import Organization
from app.db.session import get_db
from app.main import create_app

ORG = uuid.UUID("00000000-0000-0000-0000-000000000001")


@pytest_asyncio.fixture
async def client(db_engine):
    factory = async_sessionmaker(bind=db_engine, class_=AsyncSession, expire_on_commit=False)
    async with factory() as s:
        s.add(Organization(id=ORG, name="X", slug="x", country_code="HR", base_currency="EUR"))
        await s.commit()
    app = create_app()

    async def override_db():
        async with factory() as session:
            try:
                yield session
                await session.commit()
            except Exception:
                await session.rollback()
                raise

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_current_org_id] = lambda: ORG
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as ac:
        yield ac


def _xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["SKU", "Naziv", "Kol", "Cijena"])
    ws.append(["LED-6060", "LED panel 60x60 40W", "100", "30.00"])
    ws.append(["KBL-15", "Kabel NYM 3x1.5", "500", "1.20"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdf() -> bytes:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 10)
    for n, w in [("SKU", 40), ("Naziv", 80), ("Kol", 25), ("Cijena", 30)]:
        pdf.cell(w, 8, n, border=1)
    pdf.ln()
    pdf.set_font("Helvetica", size=9)
    for r in [("SPOT-7", "Spot GU10 7W", "50", "4.20")]:
        for (_n, w), v in zip([("", 40), ("", 80), ("", 25), ("", 30)], r, strict=False):
            pdf.cell(w, 8, v, border=1)
        pdf.ln()
    return bytes(pdf.output())


@pytest.mark.asyncio
async def test_import_stock_xlsx(client):
    r = await client.post("/api/v1/stock-items/import-file",
                          files={"file": ("skladiste.xlsx", _xlsx(),
                                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 201, r.text
    assert r.json()["inserted"] == 2
    items = (await client.get("/api/v1/stock-items/?search=LED")).json()["items"]
    led = next(i for i in items if i["sku"] == "LED-6060")
    assert float(led["quantity_on_hand"]) == 100
    assert float(led["unit_cost"]) == 30.0


@pytest.mark.asyncio
async def test_import_stock_pdf(client):
    r = await client.post("/api/v1/stock-items/import-file",
                          files={"file": ("skladiste.pdf", _pdf(), "application/pdf")})
    assert r.status_code == 201, r.text
    assert r.json()["inserted"] == 1


@pytest.mark.asyncio
async def test_import_stock_unsupported(client):
    r = await client.post("/api/v1/stock-items/import-file",
                          files={"file": ("x.txt", b"hello", "text/plain")})
    assert r.status_code == 415
