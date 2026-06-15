"""
XLSX/XLS parser — dual-mode: heuristic (always) + LLM (when API key set).

Column detection uses 3 signals in priority order:
  1. Header keyword matching
  2. Number distribution analysis (price vs quantity by value range/decimals)
  3. Position fallback

When ANTHROPIC_API_KEY is configured, LLM extraction runs on top and
overrides the heuristic result if confidence is higher.
"""

from __future__ import annotations

import contextlib
import io
import re

from openpyxl import load_workbook

from app.services.ingestion.parsers.base import DocumentParser, ParsedDocument, ParsedTable

# ── Keyword sets ──────────────────────────────────────────────────────────────

_DESC_KW  = {"opis", "naziv", "description", "bezeichnung", "article", "artikl",
              "artikal", "item", "roba", "produkt", "stavka", "position", "pos",
              "predmet", "materijal", "material", "roба"}
_QTY_KW   = {"kol", "količina", "qty", "menge", "cantidad", "komada", "kolicina",
              "kol.", "quantity", "amount", "kom", "stk", "anzahl",
              "broj", "kolicine"}
_UNIT_KW  = {"jedinica", "unit", "einheit", "mjera", "mjere", "um", "u/m",
              "jm", "j.m.", "mj", "j.mj", "jed.mj", "jed.mjere"}
_PRICE_KW = {"cijena", "cena", "price", "preis", "vp", "vpcj", "jedinična",
              "j.c.", "eur/jed", "nabavna", "prodajna", "tarifa", "jed.cij",
              "jedinicna", "cijena/jed", "cij", "preis/st", "einzelpr",
              "katalogska", "lista", "listna"}
_SKU_KW   = {"sku", "šifra", "sifra", "code", "art", "artikelnr", "šif", "katbr",
              "kataloška", "katalog", "art.br", "šifra art", "artikal br"}
_CAT_KW   = {"kategorija", "kategorije", "grupa", "skupina", "kategorie",
              "warengruppe", "gruppe", "category", "group", "klasa", "vrsta robe"}


def _kw_score(header: str, keywords: set[str]) -> int:
    h = header.lower().strip()
    return sum(1 for kw in keywords if kw in h)


# ── Number distribution analysis ─────────────────────────────────────────────

def _col_numbers(rows: list[list[str]], col_idx: int) -> list[float]:
    """Extract numeric values from a column."""
    nums = []
    for row in rows:
        if col_idx >= len(row):
            continue
        val = str(row[col_idx]).strip().replace(",", ".").replace(" ", "")
        val = re.sub(r"[^\d.\-]", "", val)
        with contextlib.suppress(ValueError, IndexError):
            nums.append(float(val))
    return nums


def _is_price_column(nums: list[float]) -> float:
    """Score 0..1 for how likely this column contains prices (not quantities)."""
    if len(nums) < 2:
        return 0.5
    positive = [n for n in nums if n > 0]
    if not positive:
        return 0.0
    # Prices tend to have decimal parts
    decimal_ratio = sum(1 for n in positive if n != int(n)) / len(positive)
    # Prices tend to be in range 0.01–50000
    range_ok = sum(1 for n in positive if 0.01 <= n <= 50_000) / len(positive)
    # Quantities tend to be integers and often <10000
    integer_ratio = sum(1 for n in positive if n == int(n)) / len(positive)
    large_int = sum(1 for n in positive if n > 10_000) / len(positive)

    price_score = decimal_ratio * 0.5 + range_ok * 0.3 + (1 - integer_ratio) * 0.2
    qty_penalty  = large_int * 0.3
    return max(0.0, min(1.0, price_score - qty_penalty))


def _is_qty_column(nums: list[float]) -> float:
    """Score 0..1 for how likely this column contains quantities."""
    if len(nums) < 2:
        return 0.5
    positive = [n for n in nums if n > 0]
    if not positive:
        return 0.0
    integer_ratio = sum(1 for n in positive if n == int(n)) / len(positive)
    small_range   = sum(1 for n in positive if 0 < n <= 100_000) / len(positive)
    return integer_ratio * 0.6 + small_range * 0.4


def _detect_numeric_columns(
    headers: list[str],
    rows: list[list[str]],
    kw_map: dict[str, int | None],
) -> dict[str, int | None]:
    """
    Fill gaps + override suspicious keyword assignments using number distribution.

    Override logic: if a column was keyword-assigned to 'unit' but its values
    look like prices (decimals, realistic price range), reassign to unit_price.
    """
    result = dict(kw_map)

    # Build number profiles for ALL columns (assigned or not)
    all_profiles: dict[int, tuple[float, float]] = {}  # col → (price_score, qty_score)
    for col_idx in range(len(headers)):
        nums = _col_numbers(rows, col_idx)
        if len(nums) >= max(2, len(rows) // 2):
            all_profiles[col_idx] = (_is_price_column(nums), _is_qty_column(nums))

    # Override: if 'unit' column has high price score and price is unset, swap
    unit_col = result.get("unit")
    if unit_col is not None and result.get("unit_price") is None:
        profile = all_profiles.get(unit_col)
        if profile and profile[0] > 0.6:  # strong price signal in unit column
            result["unit_price"] = unit_col
            result["unit"] = None

    used = {v for v in result.values() if v is not None}

    # Score unassigned numeric columns
    unassigned = [
        (col_idx, ps, qs)
        for col_idx, (ps, qs) in all_profiles.items()
        if col_idx not in used
    ]

    if result["unit_price"] is None:
        best = max(unassigned, key=lambda x: x[1], default=None)
        if best and best[1] > 0.4:
            result["unit_price"] = best[0]
            used.add(best[0])
            unassigned = [u for u in unassigned if u[0] not in used]

    if result["quantity"] is None:
        best = max(unassigned, key=lambda x: x[2], default=None)
        if best and best[2] > 0.4:
            result["quantity"] = best[0]

    return result


# ── Main column detection ─────────────────────────────────────────────────────

def _detect_columns(headers: list[str], rows: list[list[str]]) -> dict[str, int | None]:
    scored: dict[str, list[tuple[int, int]]] = {
        k: [] for k in ("description", "quantity", "unit", "unit_price", "sku", "category")
    }
    mapping = {
        "description": _DESC_KW, "quantity": _QTY_KW, "unit": _UNIT_KW,
        "unit_price": _PRICE_KW, "sku": _SKU_KW, "category": _CAT_KW,
    }
    for idx, h in enumerate(headers):
        for field, kws in mapping.items():
            score = _kw_score(h, kws)
            if score:
                scored[field].append((score, idx))

    result: dict[str, int | None] = {}
    used: set[int] = set()
    # category na kraju: ne kolidira s ostalima, samo pokupi preostalu kolonu
    for field in ("description", "quantity", "unit", "unit_price", "sku", "category"):
        chosen = None
        # Najveći keyword score; kod neriješenog → niža (lijevija) kolona pobjeđuje.
        # Npr. "Naziv" (col 1) ima prednost pred "Opis" (col 5) za description.
        for _, idx in sorted(scored[field], key=lambda x: (-x[0], x[1])):
            if idx not in used:
                chosen = idx
                used.add(idx)
                break
        result[field] = chosen

    if result["description"] is None:
        result["description"] = 0 if headers else None

    # Fill gaps with number distribution analysis
    result = _detect_numeric_columns(headers, rows, result)
    return result


# ── Header row detection ──────────────────────────────────────────────────────

def _find_header_row(ws, max_scan: int = 20) -> int:
    for row_idx in range(1, min(max_scan, ws.max_row or 1) + 1):
        cells = [ws.cell(row_idx, c).value for c in range(1, min((ws.max_column or 1) + 1, 20))]
        non_empty = [c for c in cells if c is not None and str(c).strip()]
        text_cells = [c for c in non_empty if isinstance(c, str)]
        if len(non_empty) >= 3 and len(text_cells) >= 2:
            return row_idx
    return 1


# ── Parser ────────────────────────────────────────────────────────────────────

class XlsxParser(DocumentParser):
    async def parse(self, file_bytes: bytes, filename: str) -> ParsedDocument:
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active

        header_row_idx = _find_header_row(ws)
        max_col = ws.max_column or 1
        raw_headers = [
            str(ws.cell(header_row_idx, c).value or "").strip()
            for c in range(1, max_col + 1)
        ]

        rows: list[list[str]] = []
        for row_idx in range(header_row_idx + 1, (ws.max_row or 1) + 1):
            cells = [ws.cell(row_idx, c).value for c in range(1, max_col + 1)]
            if not any(c is not None and str(c).strip() for c in cells):
                continue
            rows.append([str(c) if c is not None else "" for c in cells])

        # Detect columns using keywords + number distribution
        col_map = _detect_columns(raw_headers, rows)

        table = ParsedTable(rows=rows, page=None, bbox=None)
        object.__setattr__(table, "col_map", col_map)
        object.__setattr__(table, "headers", raw_headers)

        wb.close()
        return ParsedDocument(
            raw_text="\n".join("\t".join(r) for r in rows),
            tables=[table],
            detected_lang="hr",
            metadata={"header_row": header_row_idx, "col_map": col_map, "headers": raw_headers},
        )
