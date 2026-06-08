"""Popuni KORISNIKOV Excel predložak ponude vrijednostima iz ponude.

Korisnik uploada svoj predložak (svoj layout/format). Sustav:
  1. nađe redak headera (Opis/Količina/Cijena) — ista detekcija kao za uvoz
  2. upiše stavke ponude u redove ispod (čuva formatiranje predloška)
  3. zamijeni {{placeholdere}} (klijent, projekt, datum, ukupno, pdv...) bilo gdje

Tako vrijednosti idu UNUTAR korisnikovog predloška, ne u naš format.
"""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from app.services.ingestion.parsers.xlsx import _detect_columns, _find_header_row


def _ctx(quote: dict, project_name: str, client_name: str, org_name: str) -> dict[str, str]:
    def money(v) -> str:
        return f"{float(v or 0):,.2f}"
    return {
        "klijent": client_name or "",
        "projekt": project_name or "",
        "firma": org_name or "",
        "datum": date.today().strftime("%d.%m.%Y"),
        "valuta": quote.get("currency", "EUR"),
        "verzija": str(quote.get("version", 1)),
        "podukupno": money(quote.get("subtotal")),
        "pdv": money(quote.get("tax_total")),
        "ukupno": money(quote.get("total")),
        "uvjeti": quote.get("payment_terms") or "",
        "napomena": quote.get("notes_external") or "",
    }


def _replace_placeholders(ws, ctx: dict[str, str]) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "{{" in cell.value:
                val = cell.value
                for key, repl in ctx.items():
                    val = val.replace("{{" + key + "}}", repl).replace("{{ " + key + " }}", repl)
                cell.value = val


def fill_quote_template(
    *, template_bytes: bytes, quote: dict, project_name: str = "",
    client_name: str = "", org_name: str = "Ingenium",
) -> bytes:
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    # 1. Nađi header redak + stupce (Opis/Količina/Jed./Cijena)
    header_row = _find_header_row(ws)
    max_col = ws.max_column or 1
    headers = [str(ws.cell(header_row, c).value or "").strip() for c in range(1, max_col + 1)]
    # primjer-redak ispod headera (ako postoji) za bolju detekciju brojeva
    sample = [[str(ws.cell(header_row + 1, c).value or "") for c in range(1, max_col + 1)]]
    col = _detect_columns(headers, sample)

    desc_c = (col.get("description") or 0) + 1
    qty_c = col.get("quantity")
    unit_c = col.get("unit")
    price_c = col.get("unit_price")
    # stupac "ukupno" = prvi numerički nakon cijene koji nije qty/price (heuristika)
    total_c = None
    if price_c is not None:
        for c in range(price_c + 2, max_col + 1):
            h = headers[c - 1].lower()
            if any(k in h for k in ("ukupno", "iznos", "total", "vrijednost")):
                total_c = c
                break

    # 2. Upiši stavke u redove ispod headera
    items = sorted(quote.get("line_items", []), key=lambda x: x.get("position", 0))
    for i, li in enumerate(items):
        r = header_row + 1 + i
        ws.cell(r, desc_c, li.get("description", ""))
        if qty_c is not None:
            ws.cell(r, qty_c + 1, float(li.get("quantity") or 0))
        if unit_c is not None:
            ws.cell(r, unit_c + 1, li.get("unit", ""))
        if price_c is not None:
            ws.cell(r, price_c + 1, float(li.get("unit_price") or 0))
        if total_c is not None:
            lt = li.get("line_total")
            if lt is None:
                lt = Decimal(str(li.get("quantity") or 0)) * Decimal(str(li.get("unit_price") or 0))
            ws.cell(r, total_c, float(lt))

    # 3. Zamijeni single-value placeholdere
    _replace_placeholders(ws, _ctx(quote, project_name, client_name, org_name))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
