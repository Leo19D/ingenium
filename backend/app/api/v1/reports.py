"""Izvještaji — periodična agregacija ponuda, ishoda i nabave + Excel/PDF export."""

from __future__ import annotations

import io
from collections import Counter
from datetime import date, datetime
from uuid import UUID

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_org_id
from app.db.models.organization import Organization
from app.db.models.procurement import PurchaseOrder
from app.db.models.quote import Quote, QuoteOutcome
from app.db.session import get_db

router = APIRouter()


# ── Helpers ──────────────────────────────────────────────────────────────────

def _default_period() -> tuple[date, date]:
    today = date.today()
    return today.replace(day=1), today


def _as_date(dt) -> date | None:
    if dt is None:
        return None
    return dt.date() if isinstance(dt, datetime) else dt


async def _build_report(db: AsyncSession, org_id: UUID, d_from: date, d_to: date) -> dict:
    quotes = (await db.execute(
        select(Quote.id, Quote.status, Quote.total, Quote.margin_pct,
               Quote.created_at, Quote.sent_at, Quote.version)
        .where(Quote.org_id == org_id)
    )).all()
    outcomes_rows = (await db.execute(
        select(QuoteOutcome.quote_id, QuoteOutcome.outcome, QuoteOutcome.reason,
               QuoteOutcome.created_at)
        .join(Quote, Quote.id == QuoteOutcome.quote_id)
        .where(Quote.org_id == org_id)
    )).all()
    pos = (await db.execute(
        select(PurchaseOrder.status, PurchaseOrder.total, PurchaseOrder.created_at)
        .where(PurchaseOrder.org_id == org_id)
    )).all()

    def in_period(dt) -> bool:
        d = _as_date(dt)
        return d is not None and d_from <= d <= d_to

    # ── Ponude kreirane u periodu ──
    created = [q for q in quotes if in_period(q.created_at)]
    by_status: Counter = Counter(q.status for q in created)
    created_value = sum(float(q.total or 0) for q in created)
    sent_in_period = [q for q in quotes if in_period(q.sent_at)]

    # ── Ishodi zabilježeni u periodu ──
    outc = [o for o in outcomes_rows if in_period(o.created_at)]
    won = [o for o in outc if o.outcome == "won"]
    lost = [o for o in outc if o.outcome == "lost"]
    qtotal = {q.id: float(q.total or 0) for q in quotes}
    qmargin = {q.id: (float(q.margin_pct) if q.margin_pct is not None else None) for q in quotes}
    won_value = sum(qtotal.get(o.quote_id, 0) for o in won)
    lost_value = sum(qtotal.get(o.quote_id, 0) for o in lost)
    decided = len(won) + len(lost)
    win_rate = round(len(won) / decided * 100, 1) if decided else 0.0
    won_margins = [qmargin[o.quote_id] for o in won if qmargin.get(o.quote_id) is not None]
    avg_margin_won = round(sum(won_margins) / len(won_margins) * 100, 1) if won_margins else None
    lost_reasons = Counter(o.reason for o in lost if o.reason)

    # ── Nabava u periodu ──
    pos_p = [p for p in pos if in_period(p.created_at)]
    po_received = [p for p in pos_p if p.status == "received"]
    purchase_value = sum(float(p.total or 0) for p in po_received)

    return {
        "period": {"from": d_from.isoformat(), "to": d_to.isoformat()},
        "quotes": {
            "created": len(created),
            "created_value": round(created_value, 2),
            "by_status": dict(by_status),
            "sent_in_period": len(sent_in_period),
        },
        "outcomes": {
            "won": len(won), "won_value": round(won_value, 2),
            "lost": len(lost), "lost_value": round(lost_value, 2),
            "win_rate_pct": win_rate,
            "avg_margin_won_pct": avg_margin_won,
            "top_lost_reasons": lost_reasons.most_common(5),
        },
        "procurement": {
            "po_created": len(pos_p),
            "po_received": len(po_received),
            "purchase_value": round(purchase_value, 2),
        },
    }


def _parse_period(date_from: str | None, date_to: str | None) -> tuple[date, date]:
    df, dt = _default_period()
    if date_from:
        df = date.fromisoformat(date_from)
    if date_to:
        dt = date.fromisoformat(date_to)
    return df, dt


# ── JSON ─────────────────────────────────────────────────────────────────────

class ReportResponse(BaseModel):
    period: dict
    quotes: dict
    outcomes: dict
    procurement: dict


@router.get("/summary", response_model=ReportResponse)
async def report_summary(
    date_from: str | None = Query(default=None, description="YYYY-MM-DD"),
    date_to: str | None = Query(default=None, description="YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> dict:
    df, dt = _parse_period(date_from, date_to)
    return await _build_report(db, org_id, df, dt)


# ── Excel export ─────────────────────────────────────────────────────────────

@router.get("/export/xlsx")
async def report_xlsx(
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> StreamingResponse:
    df, dt = _parse_period(date_from, date_to)
    rep = await _build_report(db, org_id, df, dt)
    org = await db.get(Organization, org_id)
    org_name = org.name if org else "Ingenium"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sažetak"
    head = Font(bold=True, size=13)
    lab = Font(bold=True)
    accent = PatternFill("solid", fgColor="1A5699")
    white = Font(bold=True, color="FFFFFF")

    ws["A1"] = f"{org_name} — Izvještaj"
    ws["A1"].font = head
    ws["A2"] = f"Razdoblje: {rep['period']['from']} — {rep['period']['to']}"

    def section(row: int, title: str, rows: list[tuple[str, object]]) -> int:
        ws.cell(row, 1, title).font = white
        ws.cell(row, 1).fill = accent
        ws.cell(row, 2).fill = accent
        row += 1
        for k, v in rows:
            ws.cell(row, 1, k).font = lab
            ws.cell(row, 2, v)
            row += 1
        return row + 1

    q, o, p = rep["quotes"], rep["outcomes"], rep["procurement"]
    r = 4
    r = section(r, "PONUDE", [
        ("Kreirano", q["created"]),
        ("Vrijednost kreiranih (€)", q["created_value"]),
        ("Poslano u razdoblju", q["sent_in_period"]),
        *[(f"  status: {k}", v) for k, v in q["by_status"].items()],
    ])
    r = section(r, "ISHODI", [
        ("Dobiveno", o["won"]),
        ("Vrijednost dobivenih (€)", o["won_value"]),
        ("Izgubljeno", o["lost"]),
        ("Vrijednost izgubljenih (€)", o["lost_value"]),
        ("Win rate (%)", o["win_rate_pct"]),
        ("Prosj. marža dobivenih (%)", o["avg_margin_won_pct"] if o["avg_margin_won_pct"] is not None else "—"),
        *[(f"  razlog gubitka: {reason}", cnt) for reason, cnt in o["top_lost_reasons"]],
    ])
    r = section(r, "NABAVA", [
        ("Narudžbenice kreirane", p["po_created"]),
        ("Zaprimljeno", p["po_received"]),
        ("Nabavna vrijednost zaprimljenog (€)", p["purchase_value"]),
    ])

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20
    for row in ws.iter_rows():
        row[1].alignment = Alignment(horizontal="right")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"izvjestaj_{rep['period']['from']}_{rep['period']['to']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── PDF export ───────────────────────────────────────────────────────────────

@router.get("/export/pdf")
async def report_pdf(
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> StreamingResponse:
    df, dt = _parse_period(date_from, date_to)
    rep = await _build_report(db, org_id, df, dt)
    org = await db.get(Organization, org_id)
    from app.services.reports.pdf import generate_report_pdf

    pdf_bytes = generate_report_pdf(
        report=rep, org_name=(org.name if org else "Ingenium"),
        brand_color=((org.settings or {}).get("brand_color") if org else None),
    )
    fname = f"izvjestaj_{rep['period']['from']}_{rep['period']['to']}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes), media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
