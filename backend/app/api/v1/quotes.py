"""Quotes — CRUD + line items + totals calculation + Excel export + historical import."""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.deps import get_current_org_id, get_current_user, require_role
from app.db.models.client import Client
from app.db.models.organization import Organization
from app.db.models.project import Project
from app.db.models.quote import Quote, QuoteLineItem, QuoteOutcome
from app.db.models.user import User
from app.db.session import get_db
from app.services.audit import log_action
from app.services.email.smtp import send_email
from app.services.tax.engine import TaxContext, TaxEngine

router = APIRouter()

# ── Approval pravila — fallback ako org nema postavke ────────────────────────
APPROVAL_THRESHOLD = Decimal("5000")
DUAL_APPROVAL_THRESHOLD = Decimal("50000")
MIN_MARGIN_PCT = Decimal("0.05")


def _approval_reason(quote: Quote, limits: dict | None = None) -> str | None:
    """Razlog zašto ponuda treba odobrenje, ili None. limits iz org postavki."""
    appr = limits["approval_threshold"] if limits else APPROVAL_THRESHOLD
    dual = limits["dual_approval_threshold"] if limits else DUAL_APPROVAL_THRESHOLD
    min_m = limits["min_margin_pct"] if limits else MIN_MARGIN_PCT
    total = quote.total or Decimal("0")
    if total >= dual:
        return f"Visok iznos (≥ {dual:,.0f}) — potrebno dvostruko odobrenje"
    if total >= appr:
        return f"Iznos ≥ {appr:,.0f} — potrebno odobrenje voditelja"
    if quote.margin_pct is not None and quote.margin_pct < min_m:
        return f"Niska marža (< {float(min_m)*100:.0f}%) — potrebno odobrenje"
    return None


# ── Schemas ──────────────────────────────────────────────────────────────────

class LineItemCreate(BaseModel):
    description: str = Field(min_length=1)
    quantity: Decimal = Field(gt=0, description="Mora biti > 0")
    unit: str = "pcs"
    unit_price: Decimal = Field(ge=0, description="Ne može biti negativna")
    unit_cost: Decimal | None = Field(default=None, ge=0)
    discount_pct: Decimal = Field(default=Decimal("0"), ge=0, le=1, description="0..1 (npr. 0.1 = 10%)")
    tax_rate: Decimal | None = Field(default=None, ge=0, le=1)
    stock_item_id: UUID | None = None  # veza na skladište za praćenje zalihe
    notes: str | None = None


class LineItemUpdate(BaseModel):
    """Parcijalni update — sva polja opcionalna, primjenjuju se samo poslana."""
    description: str | None = Field(default=None, min_length=1)
    quantity: Decimal | None = Field(default=None, gt=0)
    unit: str | None = None
    unit_price: Decimal | None = Field(default=None, ge=0)
    unit_cost: Decimal | None = Field(default=None, ge=0)
    discount_pct: Decimal | None = Field(default=None, ge=0, le=1)
    tax_rate: Decimal | None = Field(default=None, ge=0, le=1)
    stock_item_id: UUID | None = None
    notes: str | None = None


class LineItemResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: UUID
    position: int
    description: str
    quantity: Decimal
    unit: str
    unit_price: Decimal
    unit_cost: Decimal | None = None
    discount_pct: Decimal
    tax_rate: Decimal | None = None
    line_total: Decimal | None = None
    margin_pct: Decimal | None = None
    stock_item_id: UUID | None = None
    notes: str | None = None


class QuoteCreate(BaseModel):
    project_id: UUID
    currency: str = "EUR"
    valid_until: date | None = None
    payment_terms: str | None = None
    notes_internal: str | None = None
    notes_external: str | None = None


class QuoteResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: UUID
    project_id: UUID
    version: int
    status: str
    currency: str
    subtotal: Decimal | None = None
    discount_total: Decimal
    tax_total: Decimal | None = None
    total: Decimal | None = None
    margin_pct: Decimal | None = None
    valid_until: date | None = None
    payment_terms: str | None = None
    notes_internal: str | None = None
    notes_external: str | None = None
    created_at: datetime
    sent_at: datetime | None = None
    line_items: list[LineItemResponse] = []


class QuoteUpdate(BaseModel):
    status: str | None = None
    currency: str | None = None
    valid_until: date | None = None
    payment_terms: str | None = None
    notes_internal: str | None = None
    notes_external: str | None = None


# ── Helpers ───────────────────────────────────────────────────────────────────

def _recalculate(quote: Quote) -> None:
    """Deterministični izračun totala iz line itemsa. Nikad LLM."""
    subtotal = Decimal("0")
    cost_total = Decimal("0")
    tax_total = Decimal("0")

    for i, item in enumerate(sorted(quote.line_items, key=lambda x: x.position), 1):
        item.position = i
        qty = item.quantity
        price = item.unit_price
        disc = item.discount_pct or Decimal("0")
        net_price = price * (1 - disc)
        line = (qty * net_price).quantize(Decimal("0.01"))
        item.line_total = line
        subtotal += line

        if item.unit_cost:
            cost = (qty * item.unit_cost).quantize(Decimal("0.01"))
            cost_total += cost
            if line > 0:
                item.margin_pct = ((line - cost) / line).quantize(Decimal("0.0001"))

        if item.tax_rate:
            tax_total += (line * item.tax_rate).quantize(Decimal("0.01"))

    quote.subtotal = subtotal
    quote.cost_total = cost_total if cost_total else None
    quote.tax_total = tax_total
    quote.total = (subtotal + tax_total).quantize(Decimal("0.01"))
    if cost_total and subtotal > 0:
        quote.margin_pct = ((subtotal - cost_total) / subtotal).quantize(Decimal("0.0001"))


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/", response_model=list[QuoteResponse])
async def list_quotes(
    project_id: UUID | None = None,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> list[QuoteResponse]:
    q = select(Quote).where(Quote.org_id == org_id).options(selectinload(Quote.line_items))
    if project_id:
        q = q.where(Quote.project_id == project_id)
    result = await db.execute(q.order_by(Quote.created_at.desc()))
    quotes = result.scalars().all()
    return [QuoteResponse.model_validate(qt) for qt in quotes]


@router.post("/", response_model=QuoteResponse, status_code=status.HTTP_201_CREATED)
async def create_quote(
    req: QuoteCreate,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
    current_user: User = Depends(get_current_user),
) -> QuoteResponse:
    # Sljedeći version broj za ovaj projekt
    existing = await db.execute(
        select(Quote.version).where(
            Quote.project_id == req.project_id, Quote.org_id == org_id
        ).order_by(Quote.version.desc()).limit(1)
    )
    last_version = existing.scalar_one_or_none() or 0

    quote = Quote(
        org_id=org_id,
        project_id=req.project_id,
        version=last_version + 1,
        currency=req.currency,
        valid_until=req.valid_until,
        payment_terms=req.payment_terms,
        notes_internal=req.notes_internal,
        notes_external=req.notes_external,
        status="draft",
        created_by=current_user.id,
        discount_total=Decimal("0"),
    )
    db.add(quote)
    await db.commit()
    # Eager-load line_items (prazna lista) da izbjegnemo lazy-load u async kontekstu
    await db.refresh(quote, attribute_names=["line_items"])
    await log_action(db, org_id=org_id, user_id=current_user.id, action="quote.created",
                     entity_type="quote", entity_id=quote.id)
    return QuoteResponse.model_validate(quote)


@router.post("/{quote_id}/new-version", response_model=QuoteResponse, status_code=status.HTTP_201_CREATED)
async def new_version(
    quote_id: UUID,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
    current_user: User = Depends(get_current_user),
) -> QuoteResponse:
    """Kreira novu verziju ponude kopiranjem stavki iz postojeće (za pregovaranje)."""
    src = (await db.execute(
        select(Quote).where(Quote.id == quote_id, Quote.org_id == org_id)
        .options(selectinload(Quote.line_items))
    )).scalar_one_or_none()
    if not src:
        raise HTTPException(status_code=404, detail="Ponuda nije pronađena.")

    last = (await db.execute(
        select(Quote.version).where(Quote.project_id == src.project_id, Quote.org_id == org_id)
        .order_by(Quote.version.desc()).limit(1)
    )).scalar_one_or_none() or 0

    new_q = Quote(
        org_id=org_id, project_id=src.project_id, version=last + 1,
        currency=src.currency, status="draft", discount_total=Decimal("0"),
        payment_terms=src.payment_terms, valid_until=src.valid_until,
        notes_internal=src.notes_internal, notes_external=src.notes_external,
        created_by=current_user.id,
    )
    db.add(new_q)
    await db.flush()
    for li in src.line_items:
        db.add(QuoteLineItem(
            quote_id=new_q.id, position=li.position, description=li.description,
            quantity=li.quantity, unit=li.unit, unit_price=li.unit_price,
            unit_cost=li.unit_cost, discount_pct=li.discount_pct,
            tax_rate=li.tax_rate, notes=li.notes,
        ))
    await db.flush()
    await db.refresh(new_q, attribute_names=["line_items"])
    _recalculate(new_q)
    await db.commit()
    await db.refresh(new_q, attribute_names=["line_items"])
    await log_action(db, org_id=org_id, user_id=current_user.id, action="quote.new_version",
                     entity_type="quote", entity_id=new_q.id,
                     after_state={"from_version": src.version, "new_version": new_q.version})
    return QuoteResponse.model_validate(new_q)


@router.get("/item-price-history")
async def item_price_history(
    description: str | None = None,
    stock_item_id: UUID | None = None,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> dict:
    """Povijest cijene po artiklu iz PROŠLIH ponuda — uči iz konkretnih cijena.

    Match: po stock_item_id (točno) ili semantičkoj sličnosti opisa. Vraća zadnju
    cijenu, prosjek dobivenih, prosj. maržu i kratku povijest s ishodima.
    """
    if not description and not stock_item_id:
        return {"matches": 0, "history": []}

    base = (
        select(
            QuoteLineItem.description, QuoteLineItem.unit_price, QuoteLineItem.margin_pct,
            Quote.version, Quote.created_at, QuoteOutcome.outcome,
        )
        .join(Quote, Quote.id == QuoteLineItem.quote_id)
        .outerjoin(QuoteOutcome, QuoteOutcome.quote_id == Quote.id)
        .where(Quote.org_id == org_id, QuoteLineItem.unit_price > 0)
    )
    if stock_item_id:
        base = base.where(QuoteLineItem.stock_item_id == stock_item_id)
    rows = (await db.execute(base)).all()

    # Bez stock_item_id → filtriraj po sličnosti opisa (semantička normalizacija)
    if not stock_item_id and description:
        from app.services.matching.catalog_matcher import _normalize
        qn = set(_normalize(description).split())

        def _sim(d: str) -> float:
            dn = set(_normalize(d or "").split())
            return len(qn & dn) / max(len(qn), len(dn)) if qn and dn else 0.0

        rows = [r for r in rows if _sim(r.description) >= 0.6]

    history = sorted(rows, key=lambda r: r.created_at or datetime.min, reverse=True)
    if not history:
        return {"matches": 0, "history": []}

    won = [r for r in history if r.outcome == "won"]
    won_prices = [float(r.unit_price) for r in won]
    margins = [float(r.margin_pct) * 100 for r in history if r.margin_pct is not None]
    return {
        "matches": len(history),
        "last_price": float(history[0].unit_price),
        "last_outcome": history[0].outcome,
        "won_count": len(won),
        "avg_won_price": round(sum(won_prices) / len(won_prices), 2) if won_prices else None,
        "avg_margin_pct": round(sum(margins) / len(margins), 1) if margins else None,
        "history": [
            {
                "price": float(r.unit_price),
                "margin_pct": round(float(r.margin_pct) * 100, 1) if r.margin_pct is not None else None,
                "outcome": r.outcome,
                "date": r.created_at.date().isoformat() if r.created_at else None,
            }
            for r in history[:8]
        ],
    }


@router.get("/{quote_id}", response_model=QuoteResponse)
async def get_quote(
    quote_id: UUID,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> QuoteResponse:
    result = await db.execute(
        select(Quote)
        .where(Quote.id == quote_id, Quote.org_id == org_id)
        .options(selectinload(Quote.line_items))
    )
    quote = result.scalar_one_or_none()
    if not quote:
        raise HTTPException(status_code=404, detail="Ponuda nije pronađena.")
    return QuoteResponse.model_validate(quote)


@router.patch("/{quote_id}", response_model=QuoteResponse)
async def update_quote(
    quote_id: UUID,
    req: QuoteUpdate,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> QuoteResponse:
    result = await db.execute(
        select(Quote)
        .where(Quote.id == quote_id, Quote.org_id == org_id)
        .options(selectinload(Quote.line_items))
    )
    quote = result.scalar_one_or_none()
    if not quote:
        raise HTTPException(status_code=404, detail="Ponuda nije pronađena.")
    for field, value in req.model_dump(exclude_none=True).items():
        setattr(quote, field, value)
    await db.commit()
    await db.refresh(quote)
    return QuoteResponse.model_validate(quote)


@router.post("/{quote_id}/line-items", response_model=QuoteResponse, status_code=status.HTTP_201_CREATED)
async def add_line_item(
    quote_id: UUID,
    req: LineItemCreate,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> QuoteResponse:
    result = await db.execute(
        select(Quote)
        .where(Quote.id == quote_id, Quote.org_id == org_id)
        .options(selectinload(Quote.line_items))
    )
    quote = result.scalar_one_or_none()
    if not quote:
        raise HTTPException(status_code=404, detail="Ponuda nije pronađena.")

    position = max((li.position for li in quote.line_items), default=0) + 1
    item = QuoteLineItem(
        quote_id=quote.id,
        position=position,
        description=req.description,
        quantity=req.quantity,
        unit=req.unit,
        unit_price=req.unit_price,
        unit_cost=req.unit_cost,
        discount_pct=req.discount_pct,
        tax_rate=req.tax_rate,
        stock_item_id=req.stock_item_id,
        notes=req.notes,
    )
    db.add(item)
    quote.line_items.append(item)
    _recalculate(quote)
    await db.commit()
    await db.refresh(quote)
    return QuoteResponse.model_validate(quote)


@router.patch("/{quote_id}/line-items/{item_id}", response_model=QuoteResponse)
async def update_line_item(
    quote_id: UUID,
    item_id: UUID,
    req: LineItemUpdate,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> QuoteResponse:
    result = await db.execute(
        select(Quote)
        .where(Quote.id == quote_id, Quote.org_id == org_id)
        .options(selectinload(Quote.line_items))
    )
    quote = result.scalar_one_or_none()
    if not quote:
        raise HTTPException(status_code=404, detail="Ponuda nije pronađena.")
    item = next((li for li in quote.line_items if li.id == item_id), None)
    if not item:
        raise HTTPException(status_code=404, detail="Stavka nije pronađena.")
    # Primijeni samo poslana polja — ostala (npr. discount/tax) ostaju netaknuta
    for field, value in req.model_dump(exclude_unset=True).items():
        setattr(item, field, value)
    _recalculate(quote)
    await db.commit()
    await db.refresh(quote)
    return QuoteResponse.model_validate(quote)


@router.delete("/{quote_id}/line-items/{item_id}", response_model=QuoteResponse)
async def delete_line_item(
    quote_id: UUID,
    item_id: UUID,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> QuoteResponse:
    result = await db.execute(
        select(Quote)
        .where(Quote.id == quote_id, Quote.org_id == org_id)
        .options(selectinload(Quote.line_items))
    )
    quote = result.scalar_one_or_none()
    if not quote:
        raise HTTPException(status_code=404, detail="Ponuda nije pronađena.")
    item = next((li for li in quote.line_items if li.id == item_id), None)
    if not item:
        raise HTTPException(status_code=404, detail="Stavka nije pronađena.")
    quote.line_items.remove(item)
    await db.delete(item)
    _recalculate(quote)
    await db.commit()
    await db.refresh(quote)
    return QuoteResponse.model_validate(quote)


@router.delete("/{quote_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_quote(
    quote_id: UUID,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> None:
    result = await db.execute(
        select(Quote).where(Quote.id == quote_id, Quote.org_id == org_id)
    )
    quote = result.scalar_one_or_none()
    if not quote:
        raise HTTPException(status_code=404, detail="Ponuda nije pronađena.")
    await db.delete(quote)
    await db.commit()


# ── Quote outcome (won/lost) ──────────────────────────────────────────────────

class OutcomeCreate(BaseModel):
    outcome: str  # won, lost, withdrawn, expired, no_response
    reason: str | None = None
    competitor_name: str | None = None
    competitor_price: Decimal | None = None
    lessons: str | None = None


@router.post("/{quote_id}/outcome", status_code=status.HTTP_201_CREATED)
async def record_outcome(
    quote_id: UUID,
    req: OutcomeCreate,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
    current_user: User = Depends(get_current_user),
) -> dict:
    result = await db.execute(
        select(Quote).where(Quote.id == quote_id, Quote.org_id == org_id)
        .options(selectinload(Quote.line_items))
    )
    quote = result.scalar_one_or_none()
    if not quote:
        raise HTTPException(status_code=404, detail="Ponuda nije pronađena.")

    existing = await db.execute(select(QuoteOutcome).where(QuoteOutcome.quote_id == quote_id))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Ishod za ovu ponudu već postoji.")

    valid = {"won", "lost", "withdrawn", "expired", "no_response"}
    if req.outcome not in valid:
        raise HTTPException(status_code=422, detail=f"Ishod mora biti jedan od: {', '.join(valid)}")

    outcome = QuoteOutcome(
        quote_id=quote_id,
        outcome=req.outcome,
        reason=req.reason,
        competitor_name=req.competitor_name,
        competitor_price=req.competitor_price,
        lessons=req.lessons,
        recorded_by=current_user.id,
    )
    db.add(outcome)
    quote.status = "accepted" if req.outcome == "won" else "rejected" if req.outcome == "lost" else req.outcome

    # Dobivena ponuda → skini zalihu za stavke vezane na skladište
    deducted = 0
    if req.outcome == "won":
        from app.services.inventory import apply_movement
        for li in quote.line_items:
            if li.stock_item_id:
                updated = await apply_movement(
                    db, org_id=org_id, stock_item_id=li.stock_item_id,
                    delta=-(li.quantity or Decimal("0")), reason="quote_won",
                    ref_type="quote", ref_id=quote_id,
                    note=f"Ponuda V{quote.version} dobivena",
                )
                if updated:
                    deducted += 1

    await db.commit()
    await log_action(db, org_id=org_id, user_id=current_user.id, action="quote.outcome",
                     entity_type="quote", entity_id=quote_id, after_state={"outcome": req.outcome})
    return {"message": "Ishod zabilježen.", "outcome": req.outcome, "stock_deducted_lines": deducted}


# ── Quote → Excel export ──────────────────────────────────────────────────────

async def _quote_with_context(quote_id: UUID, db: AsyncSession, org_id: UUID):
    """Dohvati quote + project name + client name. Raise 404 ako ne postoji."""
    result = await db.execute(
        select(Quote)
        .where(Quote.id == quote_id, Quote.org_id == org_id)
        .options(selectinload(Quote.line_items))
    )
    quote = result.scalar_one_or_none()
    if not quote:
        raise HTTPException(status_code=404, detail="Ponuda nije pronađena.")
    proj_result = await db.execute(select(Project).where(Project.id == quote.project_id))
    project = proj_result.scalar_one_or_none()
    client_name = ""
    if project and project.client_id:
        cl_result = await db.execute(select(Client).where(Client.id == project.client_id))
        cl = cl_result.scalar_one_or_none()
        if cl:
            client_name = cl.name
    return quote, (project.name if project else ""), client_name


@router.get("/{quote_id}/export/pdf")
async def export_quote_pdf(
    quote_id: UUID,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> StreamingResponse:
    """Generiraj klijent-spreman PDF dokument ponude."""
    from app.services.quote.pdf_generator import generate_quote_pdf

    quote, project_name, client_name = await _quote_with_context(quote_id, db, org_id)

    quote_dict = {
        "version": quote.version,
        "currency": quote.currency,
        "status": quote.status,
        "subtotal": quote.subtotal,
        "tax_total": quote.tax_total,
        "total": quote.total,
        "payment_terms": quote.payment_terms,
        "valid_until": quote.valid_until,
        "notes_external": quote.notes_external,
        "line_items": [
            {
                "position": li.position,
                "description": li.description,
                "quantity": li.quantity,
                "unit": li.unit,
                "unit_price": li.unit_price,
                "line_total": li.line_total,
            }
            for li in quote.line_items
        ],
    }
    brand = await _org_brand(db, org_id)
    pdf_bytes = generate_quote_pdf(
        quote=quote_dict, project_name=project_name, client_name=client_name,
        org_name=brand["org_name"], brand_color=brand["brand_color"], pdf_footer=brand["pdf_footer"],
    )
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=ponuda-V{quote.version}.pdf"},
    )


async def _org_brand(db: AsyncSession, org_id: UUID) -> dict:
    """Org naziv + brand boja + PDF footer iz postavki (za PDF ponude)."""
    org = (await db.execute(select(Organization).where(Organization.id == org_id))).scalar_one_or_none()
    s = (org.settings or {}) if org else {}
    return {
        "org_name": org.name if org else "Ingenium",
        "brand_color": s.get("brand_color", "#1a5699"),
        "pdf_footer": s.get("pdf_footer", ""),
    }


@router.post("/{quote_id}/apply-tax", response_model=QuoteResponse)
async def apply_tax(
    quote_id: UUID,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> QuoteResponse:
    """
    Izračunaj i primijeni PDV/porez po jurisdikciji (prodavač org → kupac klijent).
    Koristi TaxEngine: domaći EU → VAT, EU B2B s VAT ID → reverse charge, izvoz → 0%.
    Postavlja tax_rate na sve stavke i ponovo računa totale.
    """
    result = await db.execute(
        select(Quote).where(Quote.id == quote_id, Quote.org_id == org_id)
        .options(selectinload(Quote.line_items))
    )
    quote = result.scalar_one_or_none()
    if not quote:
        raise HTTPException(status_code=404, detail="Ponuda nije pronađena.")

    org = (await db.execute(select(Organization).where(Organization.id == org_id))).scalar_one_or_none()
    seller_country = (org.country_code if org else "HR") or "HR"
    seller_vat = (org.settings or {}).get("vat_id") if org else None

    # Kupac iz projekta → klijenta
    buyer_country = seller_country
    buyer_vat = None
    buyer_is_business = True
    proj = (await db.execute(select(Project).where(Project.id == quote.project_id))).scalar_one_or_none()
    if proj and proj.client_id:
        cl = (await db.execute(select(Client).where(Client.id == proj.client_id))).scalar_one_or_none()
        if cl:
            buyer_country = cl.country_code or seller_country
            buyer_vat = cl.tax_id

    engine = TaxEngine()
    ctx = TaxContext(
        seller_country=seller_country, seller_vat_id=seller_vat,
        buyer_country=buyer_country, buyer_vat_id=buyer_vat,
        buyer_is_business=buyer_is_business,
    )
    # Izračunaj stopu na uzorku (stopa je ista za sve stavke; iznos po stavci)
    sample = await engine.calculate(Decimal("100"), ctx)
    rate = sample.rate

    for item in quote.line_items:
        item.tax_rate = rate
    _recalculate(quote)
    await db.commit()
    await db.refresh(quote, attribute_names=["line_items"])

    resp = QuoteResponse.model_validate(quote)
    return resp


class ApprovalStatus(BaseModel):
    needs_approval: bool
    reason: str | None
    status: str
    can_send: bool


@router.get("/{quote_id}/approval-status", response_model=ApprovalStatus)
async def approval_status(
    quote_id: UUID,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> ApprovalStatus:
    res = await db.execute(select(Quote).where(Quote.id == quote_id, Quote.org_id == org_id))
    quote = res.scalar_one_or_none()
    if not quote:
        raise HTTPException(status_code=404, detail="Ponuda nije pronađena.")
    from app.api.v1.organizations import get_org_limits
    limits = await get_org_limits(db, org_id)
    reason = _approval_reason(quote, limits)
    needs = reason is not None
    can_send = (not needs) or quote.status in ("approved", "sent", "accepted")
    return ApprovalStatus(needs_approval=needs, reason=reason, status=quote.status, can_send=can_send)


@router.post("/{quote_id}/approve", response_model=QuoteResponse)
async def approve_quote(
    quote_id: UUID,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
    current_user: User = Depends(get_current_user),
    _role: str = Depends(require_role("approver")),
) -> QuoteResponse:
    """Odobri ponudu (status → approved). Samo approver/admin/owner."""
    res = await db.execute(
        select(Quote).where(Quote.id == quote_id, Quote.org_id == org_id)
        .options(selectinload(Quote.line_items))
    )
    quote = res.scalar_one_or_none()
    if not quote:
        raise HTTPException(status_code=404, detail="Ponuda nije pronađena.")
    if quote.status in ("sent", "accepted", "rejected"):
        raise HTTPException(status_code=409, detail=f"Ponuda je već u statusu '{quote.status}'.")
    quote.status = "approved"
    quote.approved_by = current_user.id
    await db.commit()
    await db.refresh(quote, attribute_names=["line_items"])
    await log_action(db, org_id=org_id, user_id=current_user.id, action="quote.approved",
                     entity_type="quote", entity_id=quote_id)
    return QuoteResponse.model_validate(quote)


class DraftEmailRequest(BaseModel):
    draft_type: str = "cover"  # cover | follow_up | thank_you


@router.post("/{quote_id}/draft-email")
async def draft_email(
    quote_id: UUID,
    req: DraftEmailRequest,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> dict:
    """AI/template draft poruke uz ponudu. Ne šalje — samo predlaže."""
    from app.services.email_assistant import generate_email_draft

    quote, project_name, client_name = await _quote_with_context(quote_id, db, org_id)
    org = (await db.execute(select(Organization).where(Organization.id == org_id))).scalar_one_or_none()
    org_name = org.name if org else "Ingenium"

    total_str = f"{float(quote.total or 0):,.2f}"
    draft = await generate_email_draft(
        draft_type=req.draft_type,
        client_name=client_name,
        project_name=project_name or "projekt",
        total=total_str,
        currency=quote.currency,
        version=quote.version,
        org_name=org_name,
    )
    return draft


class SendQuoteRequest(BaseModel):
    to_email: str
    message: str | None = None


@router.post("/{quote_id}/send")
async def send_quote(
    quote_id: UUID,
    req: SendQuoteRequest,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> dict:
    """Generiraj PDF ponude i pošalji ga klijentu emailom. Status → sent."""
    from app.services.quote.pdf_generator import generate_quote_pdf

    quote, project_name, client_name = await _quote_with_context(quote_id, db, org_id)
    if not quote.line_items:
        raise HTTPException(status_code=422, detail="Ponuda nema stavki.")

    # Approval gate — blokiraj slanje ako treba odobrenje a nije odobreno
    from app.api.v1.organizations import get_org_limits
    limits = await get_org_limits(db, org_id)
    reason = _approval_reason(quote, limits)
    if reason and quote.status not in ("approved", "sent", "accepted"):
        raise HTTPException(
            status_code=403,
            detail=f"Ponuda treba odobrenje prije slanja. {reason}",
        )

    quote_dict = {
        "version": quote.version, "currency": quote.currency, "status": quote.status,
        "subtotal": quote.subtotal, "tax_total": quote.tax_total, "total": quote.total,
        "payment_terms": quote.payment_terms, "valid_until": quote.valid_until,
        "notes_external": quote.notes_external,
        "line_items": [
            {"position": li.position, "description": li.description, "quantity": li.quantity,
             "unit": li.unit, "unit_price": li.unit_price, "line_total": li.line_total}
            for li in quote.line_items
        ],
    }
    brand = await _org_brand(db, org_id)
    pdf_bytes = generate_quote_pdf(
        quote=quote_dict, project_name=project_name, client_name=client_name,
        org_name=brand["org_name"], brand_color=brand["brand_color"], pdf_footer=brand["pdf_footer"],
    )

    total_str = f"{float(quote.total or 0):,.2f} {quote.currency}"
    custom = f"<p style='color:#444;font-size:14px;line-height:1.6'>{req.message}</p>" if req.message else ""
    html = f"""<!DOCTYPE html><html><body style="font-family:Arial,sans-serif;background:#f5f7f5;padding:32px">
      <table width="100%" cellpadding="0" cellspacing="0"><tr><td align="center">
        <table width="540" cellpadding="0" cellspacing="0" style="background:#fff;border-radius:12px;overflow:hidden;border:1px solid #e0e4e1">
          <tr><td style="background:#1a3a2a;padding:24px 32px">
            <span style="color:#a8f4b8;font-size:18px;font-weight:700">Ingenium</span>
          </td></tr>
          <tr><td style="padding:32px">
            <h1 style="font-size:20px;color:#1a231d;margin:0 0 8px">Ponuda za {_html_escape(project_name)}</h1>
            <p style="color:#667;font-size:14px;margin:0 0 20px">Poštovani{(' ' + _html_escape(client_name)) if client_name else ''},</p>
            <p style="color:#444;font-size:14px;line-height:1.6">U privitku se nalazi naša ponuda
              <strong>V{quote.version}</strong> u iznosu od <strong>{total_str}</strong>.</p>
            {custom}
            <p style="color:#444;font-size:14px;line-height:1.6">Stojimo na raspolaganju za sva pitanja.</p>
            <div style="margin-top:24px;padding-top:16px;border-top:1px solid #eee;color:#999;font-size:12px">
              Ingenium · AI Quote &amp; Procurement Platform
            </div>
          </td></tr>
        </table>
      </td></tr></table>
    </body></html>"""

    try:
        await send_email(
            to=req.to_email,
            subject=f"Ponuda V{quote.version} — {project_name}",
            html=html,
            attachment=pdf_bytes,
            attachment_name=f"ponuda-V{quote.version}.pdf",
            attachment_mime="application/pdf",
        )
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Slanje emaila nije uspjelo: {e}") from e

    quote.status = "sent"
    quote.sent_at = datetime.now(tz=__import__("datetime").timezone.utc)
    await db.commit()
    await log_action(db, org_id=org_id, action="quote.sent",
                     entity_type="quote", entity_id=quote_id,
                     after_state={"to": req.to_email})

    return {"message": f"Ponuda poslana na {req.to_email}", "status": "sent"}


def _html_escape(s: str) -> str:
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


@router.get("/{quote_id}/export/xlsx")
async def export_quote_xlsx(
    quote_id: UUID,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> StreamingResponse:
    """Generiraj klijent-spreman Excel dokument ponude."""
    result = await db.execute(
        select(Quote)
        .where(Quote.id == quote_id, Quote.org_id == org_id)
        .options(selectinload(Quote.line_items))
    )
    quote = result.scalar_one_or_none()
    if not quote:
        raise HTTPException(status_code=404, detail="Ponuda nije pronađena.")

    # Dohvati naziv projekta i klijenta
    proj_result = await db.execute(select(Project).where(Project.id == quote.project_id))
    project = proj_result.scalar_one_or_none()
    client_name = ""
    if project and project.client_id:
        cl_result = await db.execute(select(Client).where(Client.id == project.client_id))
        cl = cl_result.scalar_one_or_none()
        if cl:
            client_name = cl.name

    wb = Workbook()
    ws = wb.active
    ws.title = "Ponuda"

    # Stilovi
    dark_fill  = PatternFill("solid", fgColor="0C1410")
    green_fill = PatternFill("solid", fgColor="1E3A2A")
    head_font  = Font(bold=True, color="A8F4B8", name="Calibri", size=11)
    sub_font   = Font(color="7A9480", name="Calibri", size=9)
    body_font  = Font(color="DDEADF", name="Calibri", size=10)
    mono_font  = Font(color="A8F4B8", name="Courier New", size=10)

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "A8F4B8"

    # Header – branding
    ws.merge_cells("A1:H1")
    ws["A1"] = "⚡  INGENIUM"
    ws["A1"].font = Font(bold=True, color="A8F4B8", name="Calibri", size=16)
    ws["A1"].fill = dark_fill
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    ws["A2"] = "AI Quote & Procurement Platform  ·  ingeniumtrade.hr"
    ws["A2"].font = sub_font
    ws["A2"].fill = dark_fill

    # Separator
    ws.merge_cells("A3:H3")
    ws["A3"].fill = PatternFill("solid", fgColor="A8F4B8")
    ws.row_dimensions[3].height = 3

    # Metadata blok
    meta = [
        ("Ponuda br.",  f"V{quote.version}"),
        ("Projekt",     project.name if project else "—"),
        ("Klijent",     client_name or "—"),
        ("Valuta",      quote.currency),
        ("Vrijedi do",  str(quote.valid_until) if quote.valid_until else "—"),
        ("Uvjeti",      quote.payment_terms or "—"),
    ]
    for i, (label, value) in enumerate(meta, 5):
        ws.cell(row=i, column=1, value=label).font = sub_font
        ws.cell(row=i, column=1).fill = dark_fill
        cell = ws.cell(row=i, column=2, value=value)
        cell.font = body_font
        cell.fill = dark_fill
        ws.row_dimensions[i].height = 18

    ws.row_dimensions[4].height = 8  # razmak

    # Line items header
    header_row = len(meta) + 6
    col_labels = ["#", "Opis", "Količina", "Jed.", "Cijena/jed.", "Popust", "Porez", "Ukupno"]
    for col, label in enumerate(col_labels, 1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.fill = green_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[header_row].height = 22

    # Line items
    items = sorted(quote.line_items, key=lambda x: x.position)
    for r, item in enumerate(items, header_row + 1):
        row_fill = PatternFill("solid", fgColor="0A100E") if r % 2 == 0 else dark_fill
        vals = [
            item.position,
            item.description,
            float(item.quantity),
            item.unit,
            float(item.unit_price),
            f"{float(item.discount_pct or 0)*100:.1f}%",
            f"{float(item.tax_rate or 0)*100:.1f}%" if item.tax_rate else "—",
            float(item.line_total or 0),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = row_fill
            cell.font = mono_font if col in (1, 3, 5, 8) else body_font
            if col == 8:  # ukupno
                cell.number_format = f'"{quote.currency}" #,##0.00'
            if col == 5:
                cell.number_format = f'"{quote.currency}" #,##0.0000'
        ws.row_dimensions[r].height = 18

    # Totals — avoid writing to merged cell ranges
    total_row = header_row + len(items) + 2
    for label, value in [
        ("Međuzbroj", quote.subtotal),
        ("Porez",     quote.tax_total),
    ]:
        for c in range(1, 6):
            ws.cell(row=total_row, column=c).fill = dark_fill
        lbl_cell = ws.cell(row=total_row, column=6, value=label)
        lbl_cell.font = sub_font
        lbl_cell.fill = dark_fill
        val_cell = ws.cell(row=total_row, column=7, value=float(value or 0))
        val_cell.font = body_font
        val_cell.fill = dark_fill
        val_cell.number_format = f'"{quote.currency}" #,##0.00'
        ws.row_dimensions[total_row].height = 18
        total_row += 1

    # Grand total row
    for c in range(1, 6):
        ws.cell(row=total_row, column=c).fill = green_fill
    gt_lbl = ws.cell(row=total_row, column=6, value="UKUPNO")
    gt_lbl.font = Font(bold=True, color="A8F4B8", name="Calibri", size=11)
    gt_lbl.fill = green_fill
    gt_val = ws.cell(row=total_row, column=7, value=float(quote.total or 0))
    gt_val.font = Font(bold=True, color="A8F4B8", name="Calibri", size=11)
    gt_val.fill = green_fill
    gt_val.number_format = f'"{quote.currency}" #,##0.00'
    ws.row_dimensions[total_row].height = 24

    # Napomene
    if quote.notes_external:
        note_row = total_row + 2
        ws.merge_cells(f"A{note_row}:H{note_row}")
        ws.cell(row=note_row, column=1, value="Napomena: " + quote.notes_external).font = sub_font
        ws.cell(row=note_row, column=1).fill = dark_fill

    # Column widths
    col_widths = [5, 45, 10, 7, 13, 9, 9, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = f"A{header_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"ponuda-V{quote.version}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Historical quotes import (won/lost trening data) ─────────────────────────

class HistoricalImportResult(BaseModel):
    imported: int
    skipped: int
    errors: list[str]


@router.post("/import-historical", response_model=HistoricalImportResult)
async def import_historical_quotes(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
    current_user: User = Depends(get_current_user),
) -> HistoricalImportResult:
    """
    Uvoz historijskih ponuda iz Excela za AI trening.
    Kolone: naziv_projekta | klijent | datum | valuta | total | marza_pct | ishod | razlog | konkurent | cijena_konkurenta
    """
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        raise HTTPException(status_code=422, detail="Neispravan Excel fajl.") from exc

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    imported = 0
    skipped = 0
    errors: list[str] = []
    OUTCOME_MAP = {"won": "won", "lost": "lost", "pobijedio": "won", "izgubio": "lost",
                   "povučeno": "withdrawn", "isteklo": "expired", "nema odgovora": "no_response"}

    for i, row in enumerate(rows, 2):
        if not row or not any(row):
            skipped += 1
            continue
        try:
            project_name = str(row[0]).strip() if row[0] else f"Historijski projekt {i}"
            total_raw = row[4]
            margin_raw = row[5]
            outcome_raw = str(row[6]).strip().lower() if len(row) > 6 and row[6] else "no_response"
            outcome = OUTCOME_MAP.get(outcome_raw, "no_response")
            total = Decimal(str(total_raw).replace(",", ".")) if total_raw else Decimal("0")
            margin = Decimal(str(margin_raw).replace(",", ".").replace("%", "")) if margin_raw else None
            if margin and margin > 1:  # pretvorba ako je kao postotak npr. 25 umjesto 0.25
                margin = margin / 100

            # Kreiraj dummy projekt
            project = Project(org_id=org_id, name=project_name, status="won" if outcome == "won" else "lost")
            db.add(project)
            await db.flush()

            # Kreiraj dummy quote
            quote = Quote(
                org_id=org_id,
                project_id=project.id,
                version=1,
                currency=str(row[3]).strip().upper() if len(row) > 3 and row[3] else "EUR",
                status="accepted" if outcome == "won" else "rejected",
                total=total,
                subtotal=total,
                margin_pct=margin,
                discount_total=Decimal("0"),
                created_by=current_user.id,
            )
            db.add(quote)
            await db.flush()

            # Kreiraj outcome zapis (ključan za AI)
            qo = QuoteOutcome(
                quote_id=quote.id,
                outcome=outcome,
                reason=str(row[7]).strip() if len(row) > 7 and row[7] else None,
                competitor_name=str(row[8]).strip() if len(row) > 8 and row[8] else None,
                competitor_price=Decimal(str(row[9]).replace(",", ".")) if len(row) > 9 and row[9] else None,
                recorded_by=current_user.id,
            )
            db.add(qo)
            imported += 1
        except Exception as e:
            errors.append(f"Red {i}: {e}")
            skipped += 1

    await db.commit()
    return HistoricalImportResult(imported=imported, skipped=skipped, errors=errors[:10])
