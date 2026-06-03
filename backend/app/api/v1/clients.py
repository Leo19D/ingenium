"""Clients API endpoints — CRUD + bulk import iz Excela."""

from __future__ import annotations

import io
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_org_id
from app.db.models.client import Client
from app.db.session import get_db
from app.schemas.client import (
    BulkImportResult,
    ClientBulkImportRequest,
    ClientCreate,
    ClientResponse,
    ClientUpdate,
)

router = APIRouter()


# Mapiranje "punih naziva" zemalja iz Excela na ISO-2 kodove
COUNTRY_NAME_TO_CODE = {
    "hrvatska": "HR", "croatia": "HR",
    "slovenija": "SI", "slovenia": "SI",
    "bosna i hercegovina": "BA", "bih": "BA", "bosnia": "BA",
    "srbija": "RS", "serbia": "RS",
    "austrija": "AT", "austria": "AT",
    "njemacka": "DE", "njemačka": "DE", "germany": "DE", "deutschland": "DE",
    "italija": "IT", "italy": "IT", "italia": "IT",
    "francuska": "FR", "france": "FR",
    "spanjolska": "ES", "španjolska": "ES", "spain": "ES",
    "nizozemska": "NL", "netherlands": "NL",
    "svedska": "SE", "švedska": "SE", "sweden": "SE",
    "rumunjska": "RO", "romania": "RO",
    "poljska": "PL", "poland": "PL",
    "madjarska": "HU", "mađarska": "HU", "hungary": "HU",
    "kina": "CN", "china": "CN",
}


def _normalize_country_code(value: str | None) -> str:
    """Excel može imati 'Hrvatska', 'HR', 'Njemačka (DE)' — sve mapira na HR/DE/..."""
    if not value:
        return "HR"
    s = str(value).strip()
    # Parens npr. "Njemačka (DE)" → DE
    if "(" in s and ")" in s:
        try:
            inner = s[s.index("(") + 1 : s.index(")")].strip().upper()
            if len(inner) == 2 and inner.isalpha():
                return inner
        except ValueError:
            pass
    # Već je ISO-2
    if len(s) == 2 and s.isalpha():
        return s.upper()
    # Puni naziv
    return COUNTRY_NAME_TO_CODE.get(s.lower(), s[:2].upper())


def _parse_int(value: str | None, default: int = 30) -> int:
    if not value:
        return default
    try:
        return int(float(str(value).strip().replace(",", ".")))
    except (ValueError, TypeError):
        return default


@router.get("/")
async def list_clients(
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
    page: int = 1,
    page_size: int = 50,
    search: str | None = None,
) -> dict:
    """Lista klijenata — paginirano + pretraga po nazivu/PDV/zemlji."""
    from app.schemas.common import paginate

    base = select(Client).where(Client.org_id == org_id)
    result = await paginate(
        db, base, page=page, page_size=page_size, search=search,
        search_columns=[Client.name, Client.tax_id, Client.country_code],
        order_by=Client.created_at.desc(),
    )
    result["items"] = [ClientResponse.model_validate(c) for c in result["items"]]
    return result


@router.get("/{client_id}", response_model=ClientResponse)
async def get_client(
    client_id: UUID,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> Client:
    result = await db.execute(
        select(Client).where(Client.id == client_id, Client.org_id == org_id)
    )
    client = result.scalar_one_or_none()
    if not client:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Klijent ne postoji")
    return client


@router.post("/", response_model=ClientResponse, status_code=status.HTTP_201_CREATED)
async def create_client(
    payload: ClientCreate,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> Client:
    client = Client(org_id=org_id, **payload.model_dump())
    db.add(client)
    try:
        await db.flush()
        await db.refresh(client)
    except IntegrityError as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Klijent s ovim OIB-om već postoji: {payload.tax_id}",
        ) from e
    return client


@router.patch("/{client_id}", response_model=ClientResponse)
async def update_client(
    client_id: UUID,
    payload: ClientUpdate,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> Client:
    result = await db.execute(
        select(Client).where(Client.id == client_id, Client.org_id == org_id)
    )
    client = result.scalar_one_or_none()
    if not client:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Klijent ne postoji")
    updates = payload.model_dump(exclude_unset=True)
    for k, v in updates.items():
        setattr(client, k, v)
    await db.flush()
    await db.refresh(client)
    return client


@router.delete("/{client_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_client(
    client_id: UUID,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> None:
    result = await db.execute(
        select(Client).where(Client.id == client_id, Client.org_id == org_id)
    )
    client = result.scalar_one_or_none()
    if not client:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Klijent ne postoji")
    await db.delete(client)


@router.get("/export/xlsx")
async def export_clients_xlsx(
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> StreamingResponse:
    """Export svih klijenata kao .xlsx fajl."""
    result = await db.execute(
        select(Client).where(Client.org_id == org_id).order_by(Client.name)
    )
    clients = list(result.scalars().all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Klijenti"

    headers = ["Naziv", "Pravni naziv", "PDV broj", "Zemlja", "Industrija",
               "Segment", "Uvjeti plaćanja (dana)", "Kreditni limit", "Bilješka"]
    header_fill = PatternFill("solid", fgColor="1E3A2A")
    header_font = Font(bold=True, color="A8F4B8")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    for row, c in enumerate(clients, 2):
        ws.cell(row=row, column=1, value=c.name)
        ws.cell(row=row, column=2, value=c.legal_name)
        ws.cell(row=row, column=3, value=c.tax_id)
        ws.cell(row=row, column=4, value=c.country_code)
        ws.cell(row=row, column=5, value=c.industry)
        ws.cell(row=row, column=6, value=c.segment)
        ws.cell(row=row, column=7, value=c.payment_terms_days)
        ws.cell(row=row, column=8, value=float(c.credit_limit) if c.credit_limit else None)
        ws.cell(row=row, column=9, value=c.notes)

    col_widths = [30, 30, 16, 8, 16, 14, 10, 14, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=klijenti.xlsx"},
    )


@router.post("/bulk", response_model=BulkImportResult, status_code=status.HTTP_201_CREATED)
async def bulk_import_clients(
    payload: ClientBulkImportRequest,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> BulkImportResult:
    """Bulk uvoz klijenata iz Excela.

    Mapira polja s relaxed schemom (naziv, vat, country, segment, payment),
    preskoči duplikate po (org_id, tax_id).
    """
    inserted = 0
    skipped = 0
    errors: list[str] = []

    for idx, item in enumerate(payload.items):
        try:
            if not item.naziv or not item.naziv.strip():
                skipped += 1
                continue
            country_code = _normalize_country_code(item.country)
            segment = (item.segment or "").lower().split("/")[0].strip() or None
            client = Client(
                org_id=org_id,
                name=item.naziv.strip(),
                tax_id=(item.vat or "").strip() or None,
                country_code=country_code,
                segment=segment,
                payment_terms_days=_parse_int(item.payment, 30),
            )
            db.add(client)
            await db.flush()
            inserted += 1
        except IntegrityError:
            await db.rollback()
            skipped += 1
        except Exception as e:
            await db.rollback()
            errors.append(f"Red {idx + 1}: {type(e).__name__}: {e}")
            skipped += 1

    return BulkImportResult(inserted=inserted, skipped=skipped, errors=errors)
