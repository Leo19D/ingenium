"""Suppliers API endpoints — CRUD + bulk import iz Excela."""

from __future__ import annotations

from uuid import UUID

import io

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_org_id
from app.api.v1.clients import _normalize_country_code, _parse_int
from app.db.models.supplier import Supplier
from app.db.session import get_db
from app.schemas.client import BulkImportResult
from app.schemas.supplier import (
    SupplierBulkImportRequest,
    SupplierCreate,
    SupplierResponse,
    SupplierUpdate,
)

router = APIRouter()


@router.get("/")
async def list_suppliers(
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
    page: int = 1,
    page_size: int = 50,
    search: str | None = None,
) -> dict:
    from app.schemas.common import paginate

    base = select(Supplier).where(Supplier.org_id == org_id)
    result = await paginate(
        db, base, page=page, page_size=page_size, search=search,
        search_columns=[Supplier.name, Supplier.country_code],
        order_by=Supplier.created_at.desc(),
    )
    result["items"] = [SupplierResponse.model_validate(s) for s in result["items"]]
    return result


@router.get("/{supplier_id}", response_model=SupplierResponse)
async def get_supplier(
    supplier_id: UUID,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> Supplier:
    result = await db.execute(
        select(Supplier).where(Supplier.id == supplier_id, Supplier.org_id == org_id)
    )
    supplier = result.scalar_one_or_none()
    if not supplier:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dobavljač ne postoji")
    return supplier


@router.post("/", response_model=SupplierResponse, status_code=status.HTTP_201_CREATED)
async def create_supplier(
    payload: SupplierCreate,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> Supplier:
    supplier = Supplier(org_id=org_id, **payload.model_dump())
    db.add(supplier)
    await db.flush()
    await db.refresh(supplier)
    return supplier


@router.patch("/{supplier_id}", response_model=SupplierResponse)
async def update_supplier(
    supplier_id: UUID,
    payload: SupplierUpdate,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> Supplier:
    result = await db.execute(
        select(Supplier).where(Supplier.id == supplier_id, Supplier.org_id == org_id)
    )
    supplier = result.scalar_one_or_none()
    if not supplier:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dobavljač ne postoji")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(supplier, k, v)
    await db.flush()
    await db.refresh(supplier)
    return supplier


@router.delete("/{supplier_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_supplier(
    supplier_id: UUID,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> None:
    result = await db.execute(
        select(Supplier).where(Supplier.id == supplier_id, Supplier.org_id == org_id)
    )
    supplier = result.scalar_one_or_none()
    if not supplier:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dobavljač ne postoji")
    await db.delete(supplier)


@router.get("/export/xlsx")
async def export_suppliers_xlsx(
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> StreamingResponse:
    result = await db.execute(
        select(Supplier).where(Supplier.org_id == org_id).order_by(Supplier.name)
    )
    suppliers = list(result.scalars().all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Dobavljači"
    headers = ["Naziv", "Zemlja", "Valuta", "Incoterms", "Lead time (dana)",
               "Rating", "On-time %", "Bilješka"]
    fill = PatternFill("solid", fgColor="1E3A2A")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = Font(bold=True, color="A8F4B8")
    for row, s in enumerate(suppliers, 2):
        ws.cell(row=row, column=1, value=s.name)
        ws.cell(row=row, column=2, value=s.country_code)
        ws.cell(row=row, column=3, value=s.currency)
        ws.cell(row=row, column=4, value=s.incoterms_default)
        ws.cell(row=row, column=5, value=s.lead_time_days_avg)
        ws.cell(row=row, column=6, value=float(s.rating) if s.rating else None)
        ws.cell(row=row, column=7, value=float(s.on_time_rate) if s.on_time_rate else None)
        ws.cell(row=row, column=8, value=s.notes)
    for i, w in enumerate([30, 10, 8, 10, 8, 8, 8, 30], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dobavljaci.xlsx"},
    )


@router.post("/bulk", response_model=BulkImportResult, status_code=status.HTTP_201_CREATED)
async def bulk_import_suppliers(
    payload: SupplierBulkImportRequest,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> BulkImportResult:
    inserted = 0
    skipped = 0
    errors: list[str] = []

    for idx, item in enumerate(payload.items):
        try:
            if not item.naziv or not item.naziv.strip():
                skipped += 1
                continue
            supplier = Supplier(
                org_id=org_id,
                name=item.naziv.strip(),
                country_code=_normalize_country_code(item.country),
                currency=(item.currency or "EUR").upper().strip(),
                incoterms_default=(item.incoterms or "EXW").upper().strip(),
                lead_time_days_avg=_parse_int(item.lead, 0) or None,
            )
            db.add(supplier)
            await db.flush()
            inserted += 1
        except IntegrityError:
            await db.rollback()
            skipped += 1
        except Exception as e:  # noqa: BLE001
            await db.rollback()
            errors.append(f"Red {idx + 1}: {type(e).__name__}: {e}")
            skipped += 1

    return BulkImportResult(inserted=inserted, skipped=skipped, errors=errors)
