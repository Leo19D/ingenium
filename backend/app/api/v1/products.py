"""Products catalog — CRUD + Excel/PDF bulk import + export."""

from __future__ import annotations

import io
from typing import Annotated
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, ConfigDict
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_org_id
from app.db.models.product import Product
from app.db.session import get_db

router = APIRouter()


class ProductCreate(BaseModel):
    sku: str
    name: str
    description: str | None = None
    category: str | None = None
    brand: str | None = None
    unit: str = "pcs"


class ProductResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: UUID
    sku: str
    name: str
    description: str | None = None
    category: str | None = None
    brand: str | None = None
    unit: str
    is_active: bool


class ProductUpdate(BaseModel):
    name: str | None = None
    description: str | None = None
    category: str | None = None
    brand: str | None = None
    unit: str | None = None
    is_active: bool | None = None


@router.get("/")
async def list_products(
    category: str | None = None,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
    page: int = 1,
    page_size: int = 50,
    search: str | None = None,
) -> dict:
    from app.schemas.common import paginate

    q = select(Product).where(Product.org_id == org_id, Product.is_active == True)
    if category:
        q = q.where(Product.category == category)
    result = await paginate(
        db, q, page=page, page_size=page_size, search=search,
        search_columns=[Product.sku, Product.name, Product.category, Product.brand],
        order_by=Product.name,
    )
    result["items"] = [ProductResponse.model_validate(p) for p in result["items"]]
    return result


@router.post("/", response_model=ProductResponse, status_code=status.HTTP_201_CREATED)
async def create_product(
    req: ProductCreate,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> Product:
    product = Product(
        org_id=org_id,
        sku=req.sku.strip(),
        name=req.name.strip(),
        description=req.description,
        category=req.category,
        brand=req.brand,
        unit=req.unit,
        is_active=True,
    )
    db.add(product)
    try:
        await db.commit()
        await db.refresh(product)
    except IntegrityError:
        await db.rollback()
        raise HTTPException(status_code=409, detail=f"SKU '{req.sku}' već postoji.") from None
    return product


@router.get("/{product_id}", response_model=ProductResponse)
async def get_product(
    product_id: UUID,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> Product:
    result = await db.execute(
        select(Product).where(Product.id == product_id, Product.org_id == org_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Artikal nije pronađen.")
    return product


@router.patch("/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: UUID,
    req: ProductUpdate,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> Product:
    result = await db.execute(
        select(Product).where(Product.id == product_id, Product.org_id == org_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Artikal nije pronađen.")
    for field, value in req.model_dump(exclude_none=True).items():
        setattr(product, field, value)
    await db.commit()
    await db.refresh(product)
    return product


class BulkProductItem(BaseModel):
    sku: str = ""
    name: str = ""
    category: str | None = None
    brand: str | None = None
    unit: str = "pcs"
    description: str | None = None


class BulkProductRequest(BaseModel):
    items: list[BulkProductItem]


class BulkResult(BaseModel):
    inserted: int
    updated: int
    skipped: int
    errors: list[str] = []


@router.post("/bulk", response_model=BulkResult, status_code=status.HTTP_201_CREATED)
async def bulk_import_products(
    payload: BulkProductRequest,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> BulkResult:
    """Bulk uvoz artikala iz Excela. Postojeći SKU → update naziva/kategorije."""
    inserted = updated = skipped = 0
    errors: list[str] = []

    # Učitaj postojeće SKU-ove za upsert
    existing_q = await db.execute(select(Product).where(Product.org_id == org_id))
    by_sku = {p.sku.lower(): p for p in existing_q.scalars().all()}

    for idx, item in enumerate(payload.items, 1):
        sku = (item.sku or "").strip()
        name = (item.name or "").strip()
        if not sku or not name:
            skipped += 1
            continue
        try:
            existing = by_sku.get(sku.lower())
            if existing:
                existing.name = name
                existing.category = item.category or existing.category
                existing.brand = item.brand or existing.brand
                existing.unit = item.unit or existing.unit
                existing.description = item.description or existing.description
                existing.is_active = True
                updated += 1
            else:
                p = Product(
                    org_id=org_id, sku=sku, name=name,
                    category=item.category, brand=item.brand,
                    unit=item.unit or "pcs", description=item.description,
                    is_active=True,
                )
                db.add(p)
                await db.flush()
                by_sku[sku.lower()] = p
                inserted += 1
        except IntegrityError:
            await db.rollback()
            skipped += 1
        except Exception as e:
            errors.append(f"Red {idx}: {type(e).__name__}")
            skipped += 1

    await db.commit()
    return BulkResult(inserted=inserted, updated=updated, skipped=skipped, errors=errors[:10])


@router.post("/import-file", response_model=BulkResult)
async def import_products_from_file(
    file: Annotated[UploadFile, File()],
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> BulkResult:
    """
    Uvoz kataloga iz Excel ILI PDF fajla.
    Koristi isti ingestion parser kao RFQ — detektira SKU/naziv/kategoriju kolone.
    """
    from app.services.ingestion.parsers.pdf import PdfParser
    from app.services.ingestion.parsers.xlsx import XlsxParser

    content = await file.read()
    fname = (file.filename or "").lower()

    if fname.endswith((".xlsx", ".xls")):
        parser = XlsxParser()
    elif fname.endswith(".pdf"):
        parser = PdfParser()
    else:
        raise HTTPException(status_code=415, detail="Podržani formati: XLSX, XLS, PDF.")

    try:
        parsed = await parser.parse(content, file.filename or "katalog")
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Greška pri čitanju fajla: {e}") from e

    # Izvuci stavke iz tablica
    items: list[BulkProductItem] = []
    for table in parsed.tables:
        col_map = getattr(table, "col_map", None) or {}
        desc_col = col_map.get("description", 0)
        sku_col = col_map.get("sku")
        for row in table.rows:
            def get(c, row=row):  # row=row: bind po iteraciji (izbjegni closure loop var)
                return (row[c].strip() if c is not None and c < len(row) and row[c] else "")
            name = get(desc_col)
            if not name or len(name) < 2:
                continue
            sku = get(sku_col) or _gen_sku(name)
            items.append(BulkProductItem(sku=sku, name=name))

    if not items:
        raise HTTPException(status_code=422, detail="Nije pronađena nijedna stavka u fajlu.")

    result = await bulk_import_products(BulkProductRequest(items=items), db, org_id)
    return result


def _gen_sku(name: str) -> str:
    """Generiraj SKU iz naziva kad ga dokument nema."""
    import re
    base = re.sub(r"[^A-Za-z0-9]+", "-", name.upper())[:24].strip("-")
    return base or "ART"


@router.get("/export/xlsx")
async def export_products_xlsx(
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> StreamingResponse:
    result = await db.execute(
        select(Product).where(Product.org_id == org_id).order_by(Product.name)
    )
    products = list(result.scalars().all())
    wb = Workbook()
    ws = wb.active
    ws.title = "Katalog"
    headers = ["SKU", "Naziv", "Kategorija", "Brend", "Jedinica", "Opis"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor="1E3A2A")
        cell.font = Font(bold=True, color="A8F4B8")
    for r, p in enumerate(products, 2):
        ws.cell(row=r, column=1, value=p.sku)
        ws.cell(row=r, column=2, value=p.name)
        ws.cell(row=r, column=3, value=p.category)
        ws.cell(row=r, column=4, value=p.brand)
        ws.cell(row=r, column=5, value=p.unit)
        ws.cell(row=r, column=6, value=p.description)
    for i, w in enumerate([20, 40, 16, 16, 8, 40], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=katalog.xlsx"},
    )


@router.delete("/{product_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_product(
    product_id: UUID,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> None:
    result = await db.execute(
        select(Product).where(Product.id == product_id, Product.org_id == org_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Artikal nije pronađen.")
    await db.delete(product)
    await db.commit()


# ─────────────────────────────────────────────────────────────────────────────
# Supplier price lists — više dobavljača po artiklu, usporedba
# ─────────────────────────────────────────────────────────────────────────────

class SupplierLinkRequest(BaseModel):
    supplier_id: UUID
    unit_price: float
    currency: str = "EUR"
    supplier_sku: str | None = None
    moq: int = 1
    lead_time_days: int | None = None


@router.post("/{product_id}/suppliers", status_code=status.HTTP_201_CREATED)
async def add_supplier_to_product(
    product_id: UUID,
    req: SupplierLinkRequest,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> dict:
    """Poveži dobavljača s artiklom + cijena (ide u price history)."""
    from datetime import UTC, datetime
    from decimal import Decimal

    from app.db.models.product import SupplierPriceHistory, SupplierProduct
    from app.db.models.supplier import Supplier

    prod = (await db.execute(select(Product).where(Product.id == product_id, Product.org_id == org_id))).scalar_one_or_none()
    if not prod:
        raise HTTPException(status_code=404, detail="Artikal nije pronađen.")
    sup = (await db.execute(select(Supplier).where(Supplier.id == req.supplier_id, Supplier.org_id == org_id))).scalar_one_or_none()
    if not sup:
        raise HTTPException(status_code=404, detail="Dobavljač nije pronađen.")

    # Postojeći link ili novi
    sp = (await db.execute(
        select(SupplierProduct).where(
            SupplierProduct.product_id == product_id, SupplierProduct.supplier_id == req.supplier_id
        )
    )).scalar_one_or_none()
    if not sp:
        sp = SupplierProduct(
            product_id=product_id, supplier_id=req.supplier_id,
            supplier_sku=req.supplier_sku, supplier_name=sup.name,
            moq=req.moq, lead_time_days=req.lead_time_days, is_active=True,
        )
        db.add(sp)
        await db.flush()
    else:
        sp.moq = req.moq
        sp.lead_time_days = req.lead_time_days
        if req.supplier_sku:
            sp.supplier_sku = req.supplier_sku
        sp.is_active = True

    # Nova cijena u price history (append-only)
    db.add(SupplierPriceHistory(
        supplier_product_id=sp.id,
        unit_price=Decimal(str(req.unit_price)),
        currency=req.currency.upper(),
        valid_from=datetime.now(UTC),
        source="manual",
    ))
    await db.commit()
    return {"message": f"{sup.name} povezan s artiklom ({req.unit_price} {req.currency.upper()}).", "supplier_product_id": str(sp.id)}


@router.get("/{product_id}/suppliers")
async def compare_suppliers(
    product_id: UUID,
    quote_currency: str = "EUR",
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> list[dict]:
    """
    Usporedba dobavljača za artikl: zadnja cijena, lead time, MOQ.
    Cijene se konvertiraju u quote_currency za fer usporedbu; jeftiniji prvi.
    """
    from app.db.models.product import SupplierPriceHistory, SupplierProduct
    from app.db.models.supplier import Supplier
    from app.services.fx.rates import convert

    prod = (await db.execute(select(Product).where(Product.id == product_id, Product.org_id == org_id))).scalar_one_or_none()
    if not prod:
        raise HTTPException(status_code=404, detail="Artikal nije pronađen.")

    sps = list((await db.execute(
        select(SupplierProduct).where(SupplierProduct.product_id == product_id, SupplierProduct.is_active == True)
    )).scalars().all())
    if not sps:
        return []

    sp_ids = [sp.id for sp in sps]
    supplier_ids = list({sp.supplier_id for sp in sps})

    # 1 query: svi dobavljači odjednom → dict lookup
    sup_rows = (await db.execute(select(Supplier).where(Supplier.id.in_(supplier_ids)))).scalars().all()
    suppliers = {s.id: s for s in sup_rows}

    # 1 query: sve cijene za sve sp_ids, sortirano novije prvo → uzmi prvu po sp_id
    price_rows = (await db.execute(
        select(SupplierPriceHistory)
        .where(SupplierPriceHistory.supplier_product_id.in_(sp_ids))
        .order_by(SupplierPriceHistory.valid_from.desc())
    )).scalars().all()
    latest_price: dict = {}
    for pr in price_rows:
        if pr.supplier_product_id not in latest_price:  # prvi = najnoviji (već sortirano)
            latest_price[pr.supplier_product_id] = pr

    out = []
    for sp in sps:
        last = latest_price.get(sp.id)
        if not last:
            continue
        sup = suppliers.get(sp.supplier_id)
        conv = await convert(last.unit_price, last.currency, quote_currency)
        out.append({
            "supplier_id": str(sp.supplier_id),
            "supplier_name": sup.name if sup else (sp.supplier_name or "—"),
            "supplier_sku": sp.supplier_sku,
            "unit_price": float(last.unit_price),
            "currency": last.currency,
            "price_in_quote_ccy": float(conv["amount"]),
            "quote_currency": quote_currency.upper(),
            "fx_rate": float(conv["rate"]),
            "moq": sp.moq,
            "lead_time_days": sp.lead_time_days,
        })

    out.sort(key=lambda x: x["price_in_quote_ccy"])
    if out:
        out[0]["is_best"] = True
    return out
