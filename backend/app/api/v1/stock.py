"""Stock items API — CRUD + bulk import iz Excela (skladište)."""

from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Annotated
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_org_id, get_current_user, require_role
from app.db.models.procurement import StockMovement
from app.db.models.stock import StockItem, StockLocation
from app.db.models.user import User
from app.db.session import get_db
from app.schemas.client import BulkImportResult
from app.schemas.stock import (
    StockBulkImportItem,
    StockBulkImportRequest,
    StockItemCreate,
    StockItemResponse,
    StockItemUpdate,
    StockLocationResponse,
)
from app.services.inventory import apply_movement

router = APIRouter()


def _parse_decimal(value: str | None, default: Decimal = Decimal("0")) -> Decimal:
    """Parsiraj broj iz Excela — handluje '1.234,56' i '1234.56'."""
    if value is None or value == "":
        return default
    try:
        s = str(value).strip().replace(" ", "")
        if "," in s and "." in s:
            if s.rindex(",") > s.rindex("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif "," in s:
            s = s.replace(",", ".")
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return default


async def _get_or_create_default_location(
    db: AsyncSession, org_id: UUID, name: str = "Skladište RI"
) -> UUID:
    result = await db.execute(
        select(StockLocation).where(
            StockLocation.org_id == org_id,
            StockLocation.name == name,
        )
    )
    location = result.scalar_one_or_none()
    if location:
        return location.id
    location = StockLocation(org_id=org_id, name=name, country_code="HR")
    db.add(location)
    await db.flush()
    return location.id


@router.get("/locations", response_model=list[StockLocationResponse])
async def list_locations(
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> list[StockLocation]:
    result = await db.execute(
        select(StockLocation).where(StockLocation.org_id == org_id).order_by(StockLocation.name)
    )
    return list(result.scalars().all())


@router.get("/")
async def list_stock_items(
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
    page: int = 1,
    page_size: int = 50,
    search: str | None = None,
) -> dict:
    from app.schemas.common import paginate

    base = select(StockItem).where(StockItem.org_id == org_id)
    result = await paginate(
        db, base, page=page, page_size=page_size, search=search,
        search_columns=[StockItem.sku, StockItem.name, StockItem.category],
        order_by=StockItem.created_at.desc(),
    )
    result["items"] = [StockItemResponse.model_validate(s) for s in result["items"]]
    return result


@router.post("/import-file", response_model=BulkImportResult, status_code=status.HTTP_201_CREATED)
async def import_stock_from_file(
    file: Annotated[UploadFile, File()],
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> BulkImportResult:
    """Uvoz skladišta iz Excel ILI PDF (uklj. skenirani — OCR). Isti pipeline kao RFQ."""
    from app.services.ingestion.parsers.pdf import PdfParser
    from app.services.ingestion.parsers.xlsx import XlsxParser

    content = await file.read()
    fname = (file.filename or "").lower()
    if fname.endswith((".xlsx", ".xls")):
        parser = XlsxParser()
    elif fname.endswith(".pdf"):
        parser = PdfParser()
    else:
        raise HTTPException(status_code=415, detail="Podržani formati: XLSX, XLS, PDF.")
    try:
        parsed = await parser.parse(content, file.filename or "skladiste")
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Greška pri čitanju fajla: {e}") from e

    items: list[StockBulkImportItem] = []
    for table in parsed.tables:
        col_map = getattr(table, "col_map", None) or {}
        desc_col = col_map.get("description", 0)
        sku_col = col_map.get("sku")
        qty_col = col_map.get("quantity")
        unit_col = col_map.get("unit")
        price_col = col_map.get("unit_price")
        for row in table.rows:
            def get(c, row=row):
                return (row[c].strip() if c is not None and c < len(row) and row[c] else "")
            name = get(desc_col)
            if not name or len(name) < 2:
                continue
            sku = get(sku_col) or _gen_stock_sku(name)
            items.append(StockBulkImportItem(
                sku=sku, naziv=name, qty=get(qty_col) or None,
                unit=get(unit_col) or None, price=get(price_col) or None,
            ))
    if not items:
        raise HTTPException(status_code=422, detail="Nije pronađena nijedna stavka u fajlu.")
    return await bulk_import_stock_items(StockBulkImportRequest(items=items), db, org_id)


def _gen_stock_sku(name: str) -> str:
    import re
    base = re.sub(r"[^A-Za-z0-9]+", "-", name.upper())[:24].strip("-")
    return base or "ART"


@router.get("/{item_id}", response_model=StockItemResponse)
async def get_stock_item(
    item_id: UUID,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> StockItem:
    result = await db.execute(
        select(StockItem).where(StockItem.id == item_id, StockItem.org_id == org_id)
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Artikl ne postoji")
    return item


# ── Kretanja zalihe + ručna korekcija ────────────────────────────────────────

class MovementResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: UUID
    delta: Decimal
    reason: str
    ref_type: str | None = None
    ref_id: UUID | None = None
    note: str | None = None
    created_at: datetime


class StockAdjustRequest(BaseModel):
    delta: Decimal = Field(description="+ ulaz / − izlaz; ne smije biti 0")
    note: str | None = None


@router.get("/{item_id}/movements", response_model=list[MovementResponse])
async def list_movements(
    item_id: UUID,
    limit: int = 50,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> list[StockMovement]:
    """Povijest kretanja zalihe za artikl (najnovije prvo)."""
    rows = await db.execute(
        select(StockMovement)
        .where(StockMovement.stock_item_id == item_id, StockMovement.org_id == org_id)
        .order_by(StockMovement.created_at.desc())
        .limit(min(limit, 200))
    )
    return list(rows.scalars().all())


@router.post("/{item_id}/adjust", response_model=StockItemResponse)
async def adjust_stock(
    item_id: UUID,
    req: StockAdjustRequest,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
    current_user: User = Depends(get_current_user),
    _: str = Depends(require_role("procurement")),
) -> StockItem:
    """Ručna korekcija zalihe (+/−) uz obavezni zapis kretanja."""
    if req.delta == 0:
        raise HTTPException(status_code=422, detail="Korekcija ne smije biti 0.")
    item = await apply_movement(
        db, org_id=org_id, stock_item_id=item_id, delta=req.delta,
        reason="manual", ref_type="user", ref_id=current_user.id,
        note=req.note or "Ručna korekcija",
    )
    if not item:
        raise HTTPException(status_code=404, detail="Artikl ne postoji")
    await db.commit()
    await db.refresh(item)
    return item


@router.post("/", response_model=StockItemResponse, status_code=status.HTTP_201_CREATED)
async def create_stock_item(
    payload: StockItemCreate,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> StockItem:
    data = payload.model_dump()
    if data.get("location_id") is None:
        data["location_id"] = await _get_or_create_default_location(db, org_id)
    item = StockItem(org_id=org_id, **data)
    db.add(item)
    try:
        await db.flush()
        await db.refresh(item)
    except IntegrityError as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Artikl s SKU '{payload.sku}' već postoji u toj lokaciji",
        ) from e
    return item


@router.patch("/{item_id}", response_model=StockItemResponse)
async def update_stock_item(
    item_id: UUID,
    payload: StockItemUpdate,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> StockItem:
    result = await db.execute(
        select(StockItem).where(StockItem.id == item_id, StockItem.org_id == org_id)
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Artikl ne postoji")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(item, k, v)
    await db.flush()
    await db.refresh(item)
    return item


@router.delete("/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_stock_item(
    item_id: UUID,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> None:
    result = await db.execute(
        select(StockItem).where(StockItem.id == item_id, StockItem.org_id == org_id)
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Artikl ne postoji")
    await db.delete(item)


@router.get("/export/xlsx")
async def export_stock_xlsx(
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> StreamingResponse:
    result = await db.execute(
        select(StockItem).where(StockItem.org_id == org_id).order_by(StockItem.sku)
    )
    items = list(result.scalars().all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Skladište"
    headers = ["SKU", "Naziv", "Kategorija", "Količina", "Jed. mjere",
               "Lokacija", "Min. zaliha", "Cijena (EUR)"]
    fill = PatternFill("solid", fgColor="1E3A2A")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = Font(bold=True, color="A8F4B8")
    for row, item in enumerate(items, 2):
        ws.cell(row=row, column=1, value=item.sku)
        ws.cell(row=row, column=2, value=item.name)
        ws.cell(row=row, column=3, value=item.category)
        ws.cell(row=row, column=4, value=float(item.quantity_on_hand) if item.quantity_on_hand else 0)
        ws.cell(row=row, column=5, value=item.unit)
        ws.cell(row=row, column=6, value=item.notes)
        ws.cell(row=row, column=7, value=float(item.min_stock_level) if item.min_stock_level else None)
        ws.cell(row=row, column=8, value=float(item.unit_cost) if item.unit_cost else None)
    for i, w in enumerate([16, 32, 14, 10, 8, 16, 10, 12], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=skladiste.xlsx"},
    )


@router.post("/bulk", response_model=BulkImportResult, status_code=status.HTTP_201_CREATED)
async def bulk_import_stock_items(
    payload: StockBulkImportRequest,
    db: AsyncSession = Depends(get_db),
    org_id: UUID = Depends(get_current_org_id),
) -> BulkImportResult:
    """Bulk uvoz skladišnih artikala iz Excela."""
    default_loc_name = payload.location_name or "Skladište RI"
    default_location_id = await _get_or_create_default_location(db, org_id, default_loc_name)

    inserted = 0
    skipped = 0
    errors: list[str] = []

    for idx, item in enumerate(payload.items):
        try:
            sku = item.sku.strip()
            name = item.naziv.strip()
            if not sku or not name:
                skipped += 1
                continue

            item_location_id = default_location_id
            if item.loc and item.loc.strip() and item.loc.strip() != default_loc_name:
                item_location_id = await _get_or_create_default_location(
                    db, org_id, item.loc.strip()
                )

            unit_cost = _parse_decimal(item.price, Decimal("0"))
            stock = StockItem(
                org_id=org_id,
                location_id=item_location_id,
                sku=sku,
                name=name,
                category=(item.cat or "").strip() or None,
                unit=(item.unit or "pcs").strip(),
                quantity_on_hand=_parse_decimal(item.qty, Decimal("0")),
                min_stock_level=_parse_decimal(item.min, Decimal("0")),
                unit_cost=unit_cost if unit_cost > 0 else None,
                currency="EUR",
            )
            # Savepoint po redu — greška u jednom redu ne ruši prethodne (data loss fix)
            async with db.begin_nested():
                db.add(stock)
            inserted += 1
        except IntegrityError:
            skipped += 1
        except Exception as e:
            errors.append(f"Red {idx + 1}: {type(e).__name__}: {e}")
            skipped += 1

    return BulkImportResult(inserted=inserted, skipped=skipped, errors=errors)
